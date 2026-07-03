from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


CR_ID_RE = re.compile(r"[RSC][1-9][-sw]\d{6}(?:r\d{1})?", re.IGNORECASE)

# Mirrors CR_ID_RE but exposes the year as a capture group; keep in sync.
_TDOC_YEAR_GROUP_RE = re.compile(r"[RSC][1-9][-sw](\d{2})\d{4}(?:r\d{1})?$", re.IGNORECASE)


def tdoc_id_year(tdoc_id: str) -> int | None:
    """Return the 2-digit year embedded in ``tdoc_id``, or ``None``.

    Centralises the structural assumption that the year sits at offset 3-4
    inside the canonical CR_ID_RE shape, so SQL filters don't have to hard-code
    a ``substr`` offset that would silently desync if the convention changes.
    """
    match = _TDOC_YEAR_GROUP_RE.fullmatch(tdoc_id)
    if not match:
        return None
    return int(match.group(1))

# Substrings that disambiguate the real header row from a title row that
# happens to mention "tdoc" (e.g. "TDoc List — RAN5#111").
_HEADER_ROW_MARKERS = frozenset(
    {"title", "source", "type", "status", "spec", "version", "release"}
)


def _is_header_row(potential: Dict[str, int]) -> bool:
    has_tdoc = any("tdoc" in key for key in potential)
    if not has_tdoc:
        return False
    return any(marker in key for key in potential for marker in _HEADER_ROW_MARKERS)


def normalize_header(value: object) -> str:
    """Normalize a cell header value to lowercase whitespace-normalized text."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"\s+", " ", text)


def to_text(value: object) -> Optional[str]:
    """Convert a cell value to trimmed text, returning ``None`` for empty/missing.

    Distinguishes "missing" (cell was ``None``) from "empty" (cell was ``""``
    or whitespace) — both surface as ``None`` here so callers don't have to
    repeat the falsy check and ORM nullable columns stay correctly typed.
    """
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _parse_date_cell(value: object) -> date | None:
    """Parse an XLSX cell value as a ``date``.

    Handles ``datetime``/``date`` instances directly and ISO-style strings
    (``YYYY-MM-DD`` or ``YYYY/MM/DD``). Unparseable values return ``None``
    rather than raising so a single bad cell can't abort an entire sync.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = to_text(value)
    if text is None:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def pick_col(header_map: Dict[str, int], candidates: Iterable[str]) -> Optional[int]:
    """Find the first matching column index for a set of header candidates.

    Exact (case-insensitive) matches are preferred; substring matches are used
    as a fallback only when no exact match exists. This avoids ``"Type"``
    matching ``"Type of CR"`` when both columns are present in the sheet.
    """
    candidates_norm = [normalize_header(c) for c in candidates]
    # First pass: prefer exact (case-insensitive) header matches.
    for candidate in candidates_norm:
        if candidate and candidate in header_map:
            return header_map[candidate]
    # Second pass: fall back to substring containment (header contains candidate).
    for candidate in candidates_norm:
        if not candidate:
            continue
        for header, col in header_map.items():
            if candidate in header:
                return col
    return None


_DATE_FIELDS = frozenset({"reservation_date", "uploaded_date"})


def read_tdoc_sheet(xlsx_bytes: bytes) -> List[Dict[str, object]]:
    logger.debug("Reading TDoc XLSX bytes (%s bytes)", len(xlsx_bytes))
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if workbook is None or not workbook.sheetnames:
        logger.error("Could not read TDoc list sheet from XLSX file")
        raise RuntimeError("Could not read TDoc list sheet from XLSX file.")

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header_map: Dict[str, int] = {}
    for row in rows:
        potential = {normalize_header(cell): idx for idx, cell in enumerate(row) if normalize_header(cell)}
        if _is_header_row(potential):
            header_map = potential
            logger.debug("Detected header map: %s", header_map)
            break

    if not header_map:
        logger.error("Could not find header row in TDoc list sheet")
        raise RuntimeError("Could not find header row in TDoc list sheet.")

    col_tdoc = pick_col(header_map, ["Tdoc", "TDoc"])
    if col_tdoc is None:
        logger.error("Could not find TDoc column in TDoc list sheet")
        raise RuntimeError("Could not find TDoc column in TDoc list sheet.")

    mapping = {
        "title": pick_col(header_map, ["Title"]),
        "source": pick_col(header_map, ["Source"]),
        "type": pick_col(header_map, ["Type"]),
        "status": pick_col(header_map, ["Status", "TDoc Status"]),
        "reservation_date": pick_col(header_map, ["Reservation Date", "Reservation date"]),
        "uploaded_date": pick_col(header_map, ["Uploaded Date", "Uploaded"]),
        "cr_cat": pick_col(header_map, ["CR Cat", "CR category", "Cat"]),
        "is_revision_of": pick_col(header_map, ["Is revision of"]),
        "revised_to": pick_col(header_map, ["Revised to"]),
        "release": pick_col(header_map, ["Release"]),
        "spec": pick_col(header_map, ["Spec"]),
        "version": pick_col(header_map, ["Version"]),
        "related_wis": pick_col(header_map, ["Related WIs", "Work item code"]),
        "cr_num": pick_col(header_map, ["CR"]),
        "cr_pack": pick_col(header_map, ["TSG CR Pack", "TSG CR pack", "CR Pack"]),
    }

    result: List[Dict[str, object]] = []
    skipped_rows = 0
    for row_index, row in enumerate(rows, start=1):
        raw_tdoc = to_text(row[col_tdoc]) if col_tdoc < len(row) else None
        if not raw_tdoc:
            logger.debug("Skipping empty row %s", row_index)
            continue

        match = CR_ID_RE.search(raw_tdoc)
        if not match:
            logger.debug("Skipping non-TDoc row %s: %s", row_index, raw_tdoc)
            skipped_rows += 1
            continue

        record: Dict[str, object] = {"tdoc": match.group(0)}
        for key, col in mapping.items():
            if col is not None and col < len(row):
                cell_value = row[col]
                if key in _DATE_FIELDS:
                    record[key] = _parse_date_cell(cell_value)
                else:
                    record[key] = to_text(cell_value)
            else:
                record[key] = None
        result.append(record)

    if skipped_rows:
        logger.warning(
            "Skipped %s row(s) with unrecognized TDoc identifier pattern %r; "
            "consider widening CR_ID_RE if upstream conventions have changed.",
            skipped_rows,
            CR_ID_RE.pattern,
        )
    logger.info("Parsed %s TDoc records from XLSX", len(result))
    return result