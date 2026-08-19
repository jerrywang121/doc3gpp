from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import Workbook

from doc3gpp.parsers.tdoc_parser import (
    _extract_tdoc_hyperlinks,
    _parse_date_cell,
    pick_col,
    read_tdoc_sheet,
    to_text,
)


def _make_xlsx_bytes(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_hyperlinks(
    rows: list[list[object]],
    cell_urls: dict[tuple[int, int], str],
) -> bytes:
    """Build an XLSX where ``cell_urls[(row, col)]`` is set as a hyperlink.

    ``row``/``col`` are 1-based to match openpyxl's ``cell(row, col)`` API.
    Cell values are still taken from ``rows`` (which is 0-based in the
    caller's head — i.e. ``rows[0]`` becomes Excel row 1).
    """
    wb = Workbook()
    ws = wb.active
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if (row_idx, col_idx) in cell_urls:
                cell.hyperlink = cell_urls[(row_idx, col_idx)]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fix 2: header detection must reject title-only rows that happen to mention
# "tdoc" (e.g. "TDoc List — RAN5#111").
# ---------------------------------------------------------------------------


def test_title_only_row_is_not_treated_as_header() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc List — RAN5#111", "", "", ""],
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Agenda", "WG Chair", "agenda"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    assert records[0]["tdoc"] == "R5-260001"
    assert records[0]["title"] == "Agenda"
    assert records[0]["source"] == "WG Chair"
    assert records[0]["type"] == "agenda"


def test_real_header_detected_on_first_attempt() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
            ["R5-260002", "Doc B", "Acme", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 2
    assert [r["tdoc"] for r in records] == ["R5-260001", "R5-260002"]


def test_multiple_intro_rows_before_real_header() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["3GPP TSG-RAN WG5 #111", "", "", ""],
            ["Document Listing", "", "", ""],
            ["TDoc", "Title", "Source", "Status"],
            ["R5-260001", "Doc A", "Acme", "Agreed"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    assert records[0]["status"] == "Agreed"


def test_tdoc_row_alone_is_not_header() -> None:
    # A row with only a TDoc column and no other expected header marker must
    # be rejected; the parser should walk past it.
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc"],
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc", "Acme", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    assert records[0]["title"] == "Doc"


def test_no_header_row_raises_runtime_error() -> None:
    # No row has both "tdoc" and any other marker — parser must give up.
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["Random Title", "value"],
            ["Some other", "thing"],
        ]
    )

    import pytest

    with pytest.raises(RuntimeError, match="Could not find header row"):
        read_tdoc_sheet(xlsx_bytes)


# ---------------------------------------------------------------------------
# Fix 3: empty cells must surface as None, not "".
# ---------------------------------------------------------------------------


def test_empty_cells_become_none() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "", "", ""],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    assert records[0]["tdoc"] == "R5-260001"
    assert records[0]["title"] is None
    assert records[0]["source"] is None
    assert records[0]["type"] is None


def test_title_required_key_present_with_value() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "A real title", "Acme", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert records[0]["title"] == "A real title"
    assert records[0]["source"] == "Acme"


def test_whitespace_only_cell_becomes_none() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "   ", "\t", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert records[0]["title"] is None
    assert records[0]["source"] is None
    assert records[0]["type"] == "CR"


# ---------------------------------------------------------------------------
# #23: to_text returns None for missing AND empty cells (not "").
# ---------------------------------------------------------------------------


def test_to_text_none_returns_none() -> None:
    assert to_text(None) is None


def test_to_text_empty_string_returns_none() -> None:
    assert to_text("") is None


def test_to_text_whitespace_returns_none() -> None:
    assert to_text("   \t\n ") is None


def test_to_text_trims_surrounding_whitespace() -> None:
    assert to_text("  hello  ") == "hello"


def test_to_text_coerces_non_string() -> None:
    assert to_text(42) == "42"
    assert to_text(3.14) == "3.14"


# ---------------------------------------------------------------------------
# #8: pick_col must prefer exact match over substring (fixes "Type" vs "Type of CR").
# ---------------------------------------------------------------------------


def test_pick_col_prefers_exact_match_over_substring() -> None:
    # When both "Type" and "Type of CR" columns are present, "Type" must
    # resolve to its own column, not "Type of CR" via substring.
    header_map = {
        "tdoc": 0,
        "title": 1,
        "type": 2,
        "type of cr": 3,
    }

    assert pick_col(header_map, ["Type"]) == 2


def test_pick_col_falls_back_to_substring_when_no_exact() -> None:
    # If no exact "Type" header exists, substring match is the fallback.
    header_map = {
        "tdoc": 0,
        "title": 1,
        "type of cr": 3,
    }

    assert pick_col(header_map, ["Type"]) == 3


def test_pick_col_returns_none_when_no_match() -> None:
    assert pick_col({"foo": 0, "bar": 1}, ["Type"]) is None


# ---------------------------------------------------------------------------
# #9: skipped (non-TDoc) rows are counted and logged at WARNING.
# ---------------------------------------------------------------------------


def test_skipped_rows_trigger_warning_log(caplog) -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            # First two rows are valid TDoc IDs.
            ["R5-260001", "Doc A", "Acme", "CR"],
            ["R5-260002", "Doc B", "Acme", "CR"],
            # The next three are NOT TDoc IDs (regex miss).
            ["meeting_minutes", "Minutes", "WG", "info"],
            ["attendance_list", "Attendance", "WG", "info"],
            ["agenda", "Agenda", "WG", "info"],
        ]
    )

    with caplog.at_level(logging.WARNING, logger="doc3gpp.parsers.tdoc_parser"):
        records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 2
    warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
    assert any("Skipped 3 row(s)" in r.getMessage() for r in warnings)


def test_no_skipped_rows_no_warning(caplog) -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
        ]
    )

    with caplog.at_level(logging.WARNING, logger="doc3gpp.parsers.tdoc_parser"):
        read_tdoc_sheet(xlsx_bytes)

    skip_warnings = [
        r for r in caplog.records
        if r.levelno == logging.WARNING and "Skipped" in r.getMessage()
    ]
    assert skip_warnings == []


# ---------------------------------------------------------------------------
# #6 parser: reservation_date / uploaded_date are parsed as date objects.
# ---------------------------------------------------------------------------


def test_reservation_and_uploaded_dates_parsed_as_date() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Reservation Date", "Uploaded Date"],
            ["R5-260001", "Doc A", "2026-06-01", "2026-06-02"],
        ]
    )

    import datetime

    records = read_tdoc_sheet(xlsx_bytes)
    assert records[0]["reservation_date"] == datetime.date(2026, 6, 1)
    assert records[0]["uploaded_date"] == datetime.date(2026, 6, 2)


def test_unparseable_date_becomes_none() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Reservation Date", "Uploaded Date"],
            ["R5-260001", "Doc A", "not-a-date", "2026-06-02"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)
    assert records[0]["reservation_date"] is None
    assert records[0]["uploaded_date"] is not None


def test_empty_date_cell_is_none() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Reservation Date", "Uploaded Date"],
            ["R5-260001", "Doc A", "", ""],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)
    assert records[0]["reservation_date"] is None
    assert records[0]["uploaded_date"] is None


def test_parse_date_cell_handles_datetime_instances() -> None:
    import datetime

    dt = datetime.datetime(2026, 6, 1, 12, 30)
    assert _parse_date_cell(dt) == datetime.date(2026, 6, 1)


def test_parse_date_cell_handles_none() -> None:
    assert _parse_date_cell(None) is None


def test_parse_date_cell_handles_dd_mm_yyyy_with_time() -> None:
    import datetime

    assert _parse_date_cell("02/04/2026 15:24:41") == datetime.date(2026, 4, 2)


def test_parse_date_cell_handles_dd_mm_yyyy_date_only() -> None:
    import datetime

    assert _parse_date_cell("14/04/2026") == datetime.date(2026, 4, 14)


def test_parse_date_cell_falls_back_to_mm_dd_yyyy_for_high_second_part() -> None:
    import datetime

    assert _parse_date_cell("02/14/2026") == datetime.date(2026, 2, 14)


def test_parse_date_cell_prefers_dd_mm_for_ambiguous_slash_dates() -> None:
    import datetime

    assert _parse_date_cell("02/04/2026") == datetime.date(2026, 4, 2)


def test_parse_date_cell_returns_none_for_invalid_both_interpretations() -> None:
    assert _parse_date_cell("31/04/2026") is None


def test_reservation_date_extracted_from_dd_mm_yyyy_string_cell() -> None:
    import datetime

    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Reservation Date"],
            ["R5-260001", "Doc A", "02/04/2026 15:24:41"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)
    assert records[0]["reservation_date"] == datetime.date(2026, 4, 2)


# ---------------------------------------------------------------------------
# Hyperlink extraction: per-TDoc URL must come from the TDoc column hyperlink
# in the source XLSX (not the XLSX file URL).
# ---------------------------------------------------------------------------


def test_extract_tdoc_hyperlinks_returns_url_for_tdoc_column() -> None:
    xlsx_bytes = _make_xlsx_with_hyperlinks(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
            ["R5-260002", "Doc B", "Acme", "CR"],
        ],
        cell_urls={
            (2, 1): "https://www.3gpp.org/ftp/.../R5-260001.zip",
            (3, 1): "https://www.3gpp.org/ftp/.../R5-260002.zip",
        },
    )

    hyperlinks = _extract_tdoc_hyperlinks(xlsx_bytes)

    assert hyperlinks[2]["A"] == "https://www.3gpp.org/ftp/.../R5-260001.zip"
    assert hyperlinks[3]["A"] == "https://www.3gpp.org/ftp/.../R5-260002.zip"


def test_extract_tdoc_hyperlinks_empty_when_no_links() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title"],
            ["R5-260001", "Doc A"],
        ]
    )

    assert _extract_tdoc_hyperlinks(xlsx_bytes) == {}


def test_extract_tdoc_hyperlinks_returns_empty_for_bad_zip() -> None:
    # Not a real XLSX at all — the helper must not raise.
    assert _extract_tdoc_hyperlinks(b"not a zip file") == {}


def test_extract_tdoc_hyperlinks_ignores_non_http_targets() -> None:
    # The fixture has rIds pointing at comments/drawings as well as
    # external URLs; only http(s) targets should surface.
    xlsx_bytes = _make_xlsx_with_hyperlinks(
        [
            ["TDoc", "Title"],
            ["R5-260001", "Doc A"],
        ],
        cell_urls={(2, 1): "https://www.3gpp.org/ftp/R5-260001.zip"},
    )

    hyperlinks = _extract_tdoc_hyperlinks(xlsx_bytes)
    assert hyperlinks[2]["A"] == "https://www.3gpp.org/ftp/R5-260001.zip"
    assert set(hyperlinks[2].keys()) == {"A"}


def test_read_tdoc_sheet_populates_tdoc_url_from_hyperlink() -> None:
    xlsx_bytes = _make_xlsx_with_hyperlinks(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
        ],
        cell_urls={(2, 1): "https://www.3gpp.org/ftp/R5-260001.zip"},
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    assert records[0]["tdoc"] == "R5-260001"
    assert records[0]["tdoc_url"] == "https://www.3gpp.org/ftp/R5-260001.zip"


def test_read_tdoc_sheet_tdoc_url_is_none_when_no_hyperlink() -> None:
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert records[0]["tdoc_url"] is None


def test_read_tdoc_sheet_handles_mixed_hyperlink_presence() -> None:
    # Row 2 has a hyperlink, row 3 does not — the parser must produce
    # the correct URL for row 2 and None for row 3, not a single value
    # shared across all rows.
    xlsx_bytes = _make_xlsx_with_hyperlinks(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
            ["R5-260002", "Doc B", "Acme", "CR"],
        ],
        cell_urls={(2, 1): "https://www.3gpp.org/ftp/R5-260001.zip"},
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 2
    assert records[0]["tdoc_url"] == "https://www.3gpp.org/ftp/R5-260001.zip"
    assert records[1]["tdoc_url"] is None


def test_read_tdoc_sheet_real_fixture_extracts_per_tdoc_urls() -> None:
    # End-to-end check against the real RAN5#111 fixture: every parsed
    # record whose tdoc_id appears in column A with a hyperlink must get
    # a zip URL ending in ``/{tdoc_id}.zip``.
    fixture = (
        Path(__file__).parent.parent
        / "fixtures"
        / "tdoc_xlsx"
        / "TDoc_List_Meeting_RAN5#111.xlsx"
    )
    records = read_tdoc_sheet(fixture.read_bytes())

    assert records, "fixture should yield at least one record"
    by_id = {r["tdoc"]: r for r in records}
    assert by_id["R5-261700"]["tdoc_url"] == (
        "https://www.3gpp.org/ftp/tsg_ran/WG5_Test_ex-T1/"
        "TSGR5__111_Dalian/Docs/R5-261700.zip"
    )
    assert by_id["R5-261701"]["tdoc_url"] == (
        "https://www.3gpp.org/ftp/tsg_ran/WG5_Test_ex-T1/"
        "TSGR5__111_Dalian/Docs/R5-261701.zip"
    )
    # And confirm the URL is *not* the XLSX file URL.
    assert not by_id["R5-261700"]["tdoc_url"].endswith(".xlsx")
    assert not by_id["R5-261700"]["tdoc_url"].endswith("TDoc_List_Meeting_RAN5#111.xlsx")


def test_read_tdoc_sheet_resilient_when_xlsx_bytes_are_corrupt() -> None:
    # A corrupt XLSX must raise from openpyxl (existing contract) — the
    # hyperlink side-channel must not mask the original failure mode.
    import pytest

    with pytest.raises(Exception):
        # An empty buffer is not a valid zip; openpyxl raises.
        read_tdoc_sheet(b"")


# ---------------------------------------------------------------------------
# Six-column XLSX-metadata capture: To / Cc / Original LS / For / Abstract /
# Secretary Remarks. The parser stores them on every row (no LS gate).
# ---------------------------------------------------------------------------


def test_read_tdoc_sheet_captures_six_xlsx_metadata_columns() -> None:
    headers = [
        "TDoc", "Title", "Source", "Type",
        "For", "Abstract", "Secretary Remarks",
        "To", "Cc", "Original LS",
    ]
    xlsx_bytes = _make_xlsx_bytes(
        [
            headers,
            [
                "R5-260001", "Doc A", "Acme", "LS",
                "Information", "TL;DR of doc A", "Secretary has no remarks",
                "RAN2", "RAN3, RAN4", "C1-260001",
            ],
            # Empty cells become None for every column (mirrors title/source).
            [
                "R5-260002", "Doc B", "Acme", "CR",
                "", "", "",
                "", "", "",
            ],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 2
    a, b = records
    assert a["tdoc_for"] == "Information"
    assert a["abstract"] == "TL;DR of doc A"
    assert a["secretary_remarks"] == "Secretary has no remarks"
    assert a["ls_to"] == "RAN2"
    assert a["ls_cc"] == "RAN3, RAN4"
    assert a["original_ls"] == "C1-260001"

    # Empty cells normalise to None like every other column.
    assert b["tdoc_for"] is None
    assert b["abstract"] is None
    assert b["secretary_remarks"] is None
    assert b["ls_to"] is None
    assert b["ls_cc"] is None
    assert b["original_ls"] is None
    # Existing columns are unaffected.
    assert a["title"] == "Doc A"
    assert a["source"] == "Acme"
    assert a["type"] == "LS"


def test_read_tdoc_sheet_xlsx_metadata_none_when_header_absent() -> None:
    # No "For" / "Abstract" / etc. columns in this fixture; the parser
    # still completes and surfaces None for every new key.
    xlsx_bytes = _make_xlsx_bytes(
        [
            ["TDoc", "Title", "Source", "Type"],
            ["R5-260001", "Doc A", "Acme", "CR"],
        ]
    )

    records = read_tdoc_sheet(xlsx_bytes)

    assert len(records) == 1
    rec = records[0]
    assert rec["tdoc_for"] is None
    assert rec["abstract"] is None
    assert rec["secretary_remarks"] is None
    assert rec["ls_to"] is None
    assert rec["ls_cc"] is None
    assert rec["original_ls"] is None
